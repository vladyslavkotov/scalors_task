from openpyxl import load_workbook
import datetime
from Position import Position

def load_sheet(path:str, sheet:str = "Sheet1")-> list[list]:
    wb = load_workbook(path)
    ws = wb[sheet]
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    return data

def group_by_date(data:list[list]) -> dict[datetime, dict[str,Position]]:
    grouped = {}
    date = data[1][0]
    for row in data[1:]:
        if date == row[0]:
            if date in grouped.keys():
                new_pos = Position(row)
                grouped[date][new_pos.ticker] = new_pos
            else:
                new_pos = Position(row)
                grouped[date] = {}
                grouped[date][new_pos.ticker] = new_pos
        else:
            new_pos = Position(row)
            date = row[0]
            grouped[date] = {}
            grouped[date][new_pos.ticker] = new_pos

    return grouped

def create_correct_quantities_start(data:dict[str,Position]) -> dict[str,int]:
    correct_quantities = {}
    for ticker, position in data.items():
        if position.qty_close != 0:
            correct_quantities[ticker] = position.qty_open + position.qty_traded_today
            if position.qty_close != position.qty_open + position.qty_traded_today:
                print(f"inconsistent quantity {ticker} when creating initial quantities")

    return correct_quantities



#count
# for x,y in grouped:
#     i = 0
#     for z in y:
#         i+=1
#     print(i)

#check if yesterday price is consistent
# for date, portfolio in grouped[1:]:
#     for ticker,position in portfolio.items():
#         if ticker in previous_day[1] and position.price_yesterday != previous_day[1][ticker].price_close:
#             print(f"{ticker} price mismatch")
#             print(f"previous day {previous_day[0]} price {previous_day[1][ticker].price_close}")
#             print(f"today {date} yesterday price {position.price_yesterday}")
#     previous_day = (date, portfolio)

#check if holding quantity is changed

# if position.qty_close != position.qty_open + position.qty_traded_today:
#     print(f"{date} {ticker} close qty != open qty + trade qty")
#
# if ticker in correct_quantities:
#     # check if holding qty changed
#     if position.qty_open != correct_quantities[ticker]:
#         print(f"{date} {ticker} holding qty changed without trade "
#               f"open qty {position.qty_open} "
#               f"correct {correct_quantities[ticker]}")
#     correct_quantities[ticker] = position.qty_close
# else:
#     correct_quantities[ticker] = position.qty_close

#check if price at close * holding qty actually matches $ value
    # if position.qty_close !=0 and position.qty_close * position.price_close != position.value:
    #     print(f"{date} {ticker} incorrect value correct {position.qty_close * position.price_close} actual {position.value}")