from openpyxl import load_workbook
import datetime
from Position import Position


class Validator:
    def __init__(self, path:str, sheet:str = "Sheet1"):
        data = self.load_sheet(path, sheet)
        grouped = list(self.group_by_date(data).items())
        self.previous_day_date, self.previous_day_portfolio = grouped[0]
        self.correct_quantities = self.create_correct_quantities_start(self.previous_day_portfolio)
        # TODO not every test requires skipping 1st day
        self.grouped = grouped[1:]
        self.threshold = 5
        self.run()


    def load_sheet(self, path: str, sheet: str) -> list[list]:
        wb = load_workbook(path)
        ws = wb[sheet]
        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        return data

    def group_by_date(self, data: list[list]) -> dict[datetime, dict[str, Position]]:
        grouped = {}
        date = data[1][0]
        for row in data[1:]:
            if date == row[0]:
                if date in grouped.keys():
                    new_pos = Position(row)
                    grouped[date][new_pos.ticker] = new_pos
                else:
                    new_pos = Position(row)
                    grouped[date] = {}
                    grouped[date][new_pos.ticker] = new_pos
            else:
                new_pos = Position(row)
                date = row[0]
                grouped[date] = {}
                grouped[date][new_pos.ticker] = new_pos

        return grouped

    def create_correct_quantities_start(self, data: dict[str, Position]) -> dict[str, int]:
        correct_quantities = {}
        for ticker, position in data.items():
            if position.qty_close != 0:
                correct_quantities[ticker] = position.qty_open + position.qty_traded_today
                if position.qty_close != position.qty_open + position.qty_traded_today:
                    print(f"inconsistent quantity {ticker} when creating initial quantities")

        return correct_quantities

    def is_qty_incorrect(self, position:Position, date:str, ticker:str):
        if position.qty_close != position.qty_open + position.qty_traded_today:
            print(f"{date} {ticker} close qty != open qty + trade qty")

    def is_holding_qty_changed(self, position:Position, date:str, ticker:str ):
        if ticker in self.correct_quantities:
            if position.qty_open != self.correct_quantities[ticker]:
                print(f"{date} {ticker} holding qty changed without trade "
                      f"open qty {position.qty_open} "
                      f"correct {self.correct_quantities[ticker]}")
            self.correct_quantities[ticker]=position.qty_close
        else:
            self.correct_quantities[ticker] = position.qty_close

    def is_value_incorrect(self, position:Position, date:str, ticker:str):
        if position.qty_close !=0 and abs(position.qty_close * position.price_close - position.value) > 0.011 :
            print(f"{date} {ticker} incorrect value calculated {position.qty_close * position.price_close} actual {position.value}")

    def is_price_equal_to_stock_mvmt(self, position:Position, date:str, ticker:str):
        calculated_price_change = round(position.price_yesterday * (1 + position.stock_movement_pct), 2)
        if abs(calculated_price_change - position.price_close) > 0.011:
            print(f"{date} {ticker} incorrect price and/or %move calculated {calculated_price_change} actual {position.price_close}")

    def is_trade_price_incorrect(self, position:Position, date:str, ticker:str):
        if (position.price_close !=0 and
            position.price_traded_today!=0 and
            abs(position.price_close - position.price_traded_today)/position.price_close > abs(position.stock_movement_pct)*self.threshold):
            print(f"{date} {ticker} price move is more than 5x of daily % move "
                  f"CALC MOVE {abs(position.price_close - position.price_traded_today)/position.price_close} "
                  f"MOVE % {position.stock_movement_pct}")

    def run(self):
        for date, portfolio in self.grouped:
            for ticker, position in portfolio.items():
                # self.is_qty_incorrect(position, date, ticker)
                # self.is_value_incorrect(position, date, ticker)
                self.is_holding_qty_changed(position, date, ticker)
                # self.is_trade_price_incorrect(position, date, ticker)
                # self.is_price_equal_to_stock_mvmt(position, date, ticker)
