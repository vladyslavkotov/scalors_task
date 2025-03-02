from Position import Position
from openpyxl import load_workbook
from datetime import datetime, date


class GroupedPositions:
    @staticmethod
    def load_sheet(path: str, sheet: str) -> list[list]:
        wb = load_workbook(path)
        ws = wb[sheet]
        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        return data

    @staticmethod
    def group_by_date(data: list[list]) -> dict[datetime, dict[str, Position]]:
        grouped = {}
        date = data[1][0].date()
        for row in data[1:]:
            if date == row[0].date():
                if date in grouped.keys():
                    new_pos = Position(row)
                    grouped[date][new_pos.ticker] = new_pos
                else:
                    new_pos = Position(row)
                    grouped[date] = {}
                    grouped[date][new_pos.ticker] = new_pos
            else:
                new_pos = Position(row)
                date = row[0].date()
                grouped[date] = {}
                grouped[date][new_pos.ticker] = new_pos

        return grouped

    @staticmethod
    def create(path: str, sheet: str = "Sheet1") -> list[tuple[datetime, dict[str, Position]]]:
        data = GroupedPositions.load_sheet(path, sheet)
        return list(GroupedPositions.group_by_date(data).items())
